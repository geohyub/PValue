"""Excel report generation."""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)


def generate_excel_report(
    results: pd.DataFrame,
    summary: pd.DataFrame,
    config: dict,
    save_path: str,
) -> bool:
    """Write a formatted Excel report with summary, full results, and task info.

    Returns:
        ``True`` on success, ``False`` if openpyxl is not installed.
    """
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        logger.warning("openpyxl is not installed — skipping Excel report")
        return False

    Path(save_path).parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        results.to_excel(writer, sheet_name="Results", index=False)

        tasks_info = pd.DataFrame(config.get("tasks", []))
        if not tasks_info.empty:
            tasks_info.to_excel(writer, sheet_name="Tasks", index=False)

        _apply_formatting(writer.book, Font, PatternFill, Alignment)

    logger.info("Excel report saved: %s", save_path)
    return True


def _apply_formatting(wb, Font, PatternFill, Alignment) -> None:
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in ws.columns:
            col_letter = column_cells[0].column_letter
            max_len = max(
                (len(str(c.value)) for c in column_cells if c.value is not None),
                default=8,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
