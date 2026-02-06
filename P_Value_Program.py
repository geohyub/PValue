# ==== Marine Pxx Simulator V5 - FINAL ====
# 주요 개선사항:
# - V4 대비 P값 결과 동일성 검증 완료
# - 한글 폰트 자동 감지 (Windows/Linux 모두 지원)
# - hindcast CSV 자동 인식 (ERA5)
# - 시간 간격 자동 처리 (10분, 1시간 등)
# - NA 처리 방식 선택 가능 (permissive/conservative)
# - work vs wait 타임라인 차트
# - 모든 P값 히스토그램 표시 (표준 색상)
# - 업무시간 제약 시각화
# - Excel 리포트 자동 생성
# - 최적 시작월 분석
# - 배치 실행 지원
# ===============================================================

from dataclasses import dataclass
from typing import Dict, List, Optional, Callable, Tuple
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import re, json, os, logging
from datetime import datetime
import platform

# ============ 한글 폰트 자동 감지 ============
def get_korean_font():
    """시스템에서 사용 가능한 한글 폰트 자동 선택"""
    fonts = [f.name for f in fm.fontManager.ttflist]
    
    # 우선순위
    priority = ["Malgun Gothic", "NanumGothic", "AppleGothic", 
                "Noto Sans KR", "NanumBarunGothic", "DejaVu Sans"]
    
    for font in priority:
        if font in fonts:
            return font
    
    # 대체: 한글 포함 폰트 검색
    for font in fonts:
        if any(k in font for k in ["Gothic", "Nanum", "Malgun", "Batang", "Dotum"]):
            return font
    
    return "DejaVu Sans"  # 최종 fallback

# matplotlib 설정
plt.rcParams["font.family"] = get_korean_font()
plt.rcParams['axes.unicode_minus'] = False

logging.basicConfig(level=logging.INFO, format='%(message)s')

# ============ 코어 데이터 구조 ============
@dataclass
class Task:
    name: str
    duration_h: int
    thresholds: Dict[str, float]
    setup_h: int = 0
    teardown_h: int = 0

# ============ 데이터 검증 ============
def validate_metocean(df: pd.DataFrame, required_cols: List[str] = ['Hs', 'Wind']) -> Tuple[bool, str]:
    """기상 데이터 검증"""
    # 1. 컬럼 확인
    for col in required_cols:
        if col not in df.columns:
            return False, f"필수 컬럼 누락: {col}"
    
    # 2. 시간 인덱스 확인
    if not isinstance(df.index, pd.DatetimeIndex):
        return False, "시간 인덱스가 DatetimeIndex가 아닙니다"
    
    # 3. 중복 확인
    if df.index.duplicated().any():
        return False, "중복된 timestamp 존재"
    
    # 4. 시간 간격 확인 (경고만, 블록하지 않음)
    intervals = df.index.to_series().diff()[1:].value_counts()
    if len(intervals) > 2:  # 불규칙하지만 계속 진행
        logging.warning(f"⚠ 시간 간격 불규칙: 가장 흔한 간격={intervals.index[0]} ({intervals.values[0]}개)")
    
    # 5. 결측치 확인
    for col in required_cols:
        na_ratio = df[col].isna().sum() / len(df) * 100
        if na_ratio > 50:
            return False, f"{col} 결측치 과다: {na_ratio:.1f}%"
    
    # 6. 값 범위 확인 (완화된 기준)
    if (df['Hs'] < 0).any() or (df['Hs'] > 20).any():
        return False, "Hs 값이 비정상 범위 (0~20m)"
    
    if (df['Wind'] < 0).any() or (df['Wind'] > 70).any():
        return False, "Wind 값이 비정상 범위 (0~70m/s)"
    
    # 7. 최소 데이터 길이
    if len(df) < 24:  # 최소 1일치
        return False, f"데이터가 너무 짧습니다: {len(df)}개"
    
    return True, "검증 통과"


def get_time_interval_minutes(df: pd.DataFrame) -> int:
    """데이터의 시간 간격을 분 단위로 반환"""
    intervals = df.index.to_series().diff()[1:].value_counts()
    most_common_interval = intervals.index[0]
    return int(most_common_interval.total_seconds() / 60)


# ============ 조건 마스크 생성 ============
def build_condition_mask(block: pd.DataFrame, thresholds: Dict[str, float],
                        na_handling: str = 'permissive') -> np.ndarray:
    """
    기상 조건 마스크 생성
    
    Args:
        block: 기상 데이터
        thresholds: 임계치 딕셔너리
        na_handling: NA 처리 방식
            - 'permissive': NA → True (작업 가능)
            - 'conservative': NA → False (작업 불가)
    """
    mask = np.ones(len(block), dtype=bool)
    
    for k, thr in thresholds.items():
        if k not in block.columns:
            raise KeyError(f"Missing column in metocean: {k}")
        
        values = block[k].values
        cond = values <= thr
        
        # NA 처리
        na_mask = np.isnan(values)
        if na_handling == 'permissive':
            cond = np.where(na_mask, True, cond)  # NA → 작업 가능
        else:  # conservative
            cond = np.where(na_mask, False, cond)  # NA → 작업 불가
        
        mask &= cond
    
    return mask


# ============ 연속 블록 방식 ============
def find_next_window(mask: np.ndarray, start_idx: int, need_steps: int) -> Tuple[int,int,int]:
    """연속 작업 가능 구간 찾기"""
    n = len(mask)
    run = 0
    start_run = None
    i = start_idx
    steps = 0
    
    while steps < 2*n:
        if mask[i]:
            if run == 0:
                start_run = i
            run += 1
            if run >= need_steps:
                end_idx = (start_run + need_steps - 1) % n
                waiting = (start_run - start_idx) % n
                return start_run, end_idx, waiting
        else:
            run = 0
            start_run = None
        i = (i + 1) % n
        steps += 1
    
    raise RuntimeError("No feasible window found")


# ============ 분할 누적 방식 ============
def find_window_accumulated(mask: np.ndarray, start_idx: int, need_steps: int) -> Tuple[int,int,int]:
    """분할 작업 가능 (누적) 방식"""
    n = len(mask)
    i = start_idx
    waited = 0
    worked = 0
    steps = 0
    started = False
    
    while steps < 2*n and worked < need_steps:
        if mask[i]:
            started = True
            worked += 1
        else:
            if not started or worked > 0:
                waited += 1
        i = (i + 1) % n
        steps += 1
    
    if worked < need_steps:
        raise RuntimeError("No feasible window found (accumulated)")
    
    end_idx = (i - 1) % n
    return end_idx, waited, worked


# ============ 캠페인 시뮬레이션 ============
def simulate_campaign(metocean: pd.DataFrame, tasks: List[Task], 
                     n_sims: int = 1000,
                     start_month: Optional[int] = None,
                     calendar_mask_fn: Optional[Callable[[pd.DatetimeIndex], np.ndarray]] = None,
                     seed: Optional[int] = 7,
                     split_mode: bool = False,
                     time_interval_min: int = 60,
                     na_handling: str = 'permissive',
                     show_progress: bool = True) -> pd.DataFrame:
    """
    몬테카를로 시뮬레이션
    
    Args:
        time_interval_min: 데이터의 시간 간격 (분)
        na_handling: NA 처리 방식 ('permissive' or 'conservative')
        show_progress: 진행 상태 표시
    """
    if seed is not None:
        np.random.seed(seed)
    
    if "year" not in metocean.columns:
        metocean = metocean.copy()
        metocean["year"] = metocean.index.year
    
    years = sorted(metocean["year"].unique())
    results = []
    
    # 시간 간격에 따른 steps 변환
    steps_per_hour = 60 / time_interval_min
    
    for sim in range(n_sims):
        if show_progress and (sim + 1) % 500 == 0:
            print(f"진행: {sim+1}/{n_sims} ({(sim+1)/n_sims*100:.1f}%)")
        
        yr = int(np.random.choice(years))
        block = metocean[metocean["year"] == yr]
        
        # 시작 시점 선택
        if start_month is None:
            start_time = block.index[0] + pd.Timedelta(hours=np.random.randint(0, len(block)))
        else:
            month_block = block[block.index.month == start_month]
            if len(month_block) == 0:
                month_block = block  # fallback
            start_time = month_block.index[np.random.randint(0, len(month_block))]
        
        start_idx = block.index.get_indexer([start_time])[0]
        current_idx = start_idx
        
        # 캘린더 마스크
        if calendar_mask_fn is not None:
            cal_mask = calendar_mask_fn(block.index)
            if len(cal_mask) != len(block):
                raise ValueError("Calendar mask length mismatch")
        else:
            cal_mask = np.ones(len(block), dtype=bool)
        
        total_elapsed_steps = 0
        total_wait_steps = 0
        
        for t in tasks:
            cond_mask = build_condition_mask(block, t.thresholds, na_handling) & cal_mask
            required_h = t.setup_h + t.duration_h + t.teardown_h
            required_steps = int(required_h * steps_per_hour)
            
            if not split_mode:
                s, e, waiting = find_next_window(cond_mask, current_idx, required_steps)
                total_wait_steps += waiting
                total_elapsed_steps += waiting + required_steps
                current_idx = (e + 1) % len(block)
            else:
                e, waiting, worked = find_window_accumulated(cond_mask, current_idx, required_steps)
                total_wait_steps += waiting
                total_elapsed_steps += waiting + worked
                current_idx = (e + 1) % len(block)
        
        # steps를 시간으로 변환
        total_elapsed_h = total_elapsed_steps / steps_per_hour
        total_wait_h = total_wait_steps / steps_per_hour
        
        results.append({
            "sim": sim,
            "year_sample": yr,
            "start_time": start_time,
            "elapsed_hours": total_elapsed_h,
            "wait_hours": total_wait_h,
            "work_hours": total_elapsed_h - total_wait_h,
            "elapsed_days": total_elapsed_h / 24.0
        })
    
    return pd.DataFrame(results)


def summarize_pxx(df: pd.DataFrame, p_list=[50,75,90]) -> pd.DataFrame:
    """P값 요약"""
    metrics = [f"P{int(p)}" for p in p_list] + ["Mean", "Std", "Min", "Max"]
    values = [np.percentile(df["elapsed_days"], p) for p in p_list]
    values += [df["elapsed_days"].mean(), df["elapsed_days"].std(),
               df["elapsed_days"].min(), df["elapsed_days"].max()]
    
    return pd.DataFrame({
        "Metric": metrics,
        "Value_days": values
    })


# ============ CSV 로더 ============
def load_csv(csv_path: str, csv_type: str = 'general', 
             start_date: Optional[str] = None, 
             end_date: Optional[str] = None) -> pd.DataFrame:
    """
    CSV 로드 (자동 인코딩 감지 및 BOM 처리)
    
    Args:
        csv_type: 'general' or 'hindcast'
        start_date: 시작 날짜 (hindcast only, 예: '1995-01-01')
        end_date: 종료 날짜 (hindcast only, 예: '2024-12-31')
    """
    # BOM 처리 및 인코딩
    encodings = ['utf-8-sig', 'utf-8', 'cp949', 'euc-kr']
    df = None
    
    for enc in encodings:
        try:
            if csv_type == 'hindcast':
                # hindcast: 5줄 스킵, 6번째 줄이 헤더
                try:
                    df_raw = pd.read_csv(csv_path, skiprows=5, encoding=enc)
                except:
                    continue
                
                # 컬럼명 정리 (대괄호 제거)
                df_raw.columns = [c.strip().replace('[', '').replace(']', '') for c in df_raw.columns]
                
                # 시간 컬럼 (첫 번째)
                time_col = df_raw.columns[0]
                
                try:
                    df_raw[time_col] = pd.to_datetime(df_raw[time_col], format='%Y-%m-%d %H:%M:%S')
                except:
                    df_raw[time_col] = pd.to_datetime(df_raw[time_col], format='%d/%m/%Y %H:%M:%S')
                
                df_raw = df_raw.set_index(time_col)
                
                # Wind10, Hs 찾기
                wind_col = None
                hs_col = None
                
                for col in df_raw.columns:
                    col_lower = col.lower()
                    if '10m' in col_lower and 'wind' in col_lower:
                        wind_col = col
                    if 'hs' in col_lower and 'm' in col_lower:
                        hs_col = col
                
                if not wind_col or not hs_col:
                    continue
                
                # Wind, Hs로 변환
                df = df_raw[[wind_col, hs_col]].copy()
                df.columns = ['Wind', 'Hs']
                df = df.apply(pd.to_numeric, errors='coerce')  # 숫자 변환
                
                # 날짜 구간 필터링
                if start_date:
                    df = df[df.index >= start_date]
                    logging.info(f"✓ 시작 날짜 필터링: {start_date}")
                if end_date:
                    df = df[df.index <= end_date]
                    logging.info(f"✓ 종료 날짜 필터링: {end_date}")
            else:
                # 일반 CSV
                df = pd.read_csv(csv_path, encoding=enc, parse_dates=['timestamp'])
                df = df.set_index('timestamp')
            
            break
        except Exception as e:
            continue
    
    if df is None:
        raise ValueError(f"CSV 로드 실패: {csv_path}")
    
    return df


# ============ 시각화 함수 ============
def get_pvalue_style(p: int, pvals: List[int]) -> Tuple[str, float, str]:
    """P값별 시각화 스타일 (표준 방식)"""
    if p == 50:
        return 'blue', 2.5, '-'      # P50: 파랑 실선
    elif p == 90:
        return 'red', 2.5, '-'       # P90: 빨강 실선 (가장 중요)
    elif p == min(pvals):
        return 'darkred', 2.0, '-'   # 최소: 진한 빨강
    elif p == max(pvals):
        return 'darkred', 2.0, ':'   # 최대: 진한 빨강 점선
    else:
        return 'gray', 1.0, '--'     # 나머지: 회색 점선


def plot_histogram_with_pvals(res: pd.DataFrame, pvals: List[int], save_path: str = None):
    """히스토그램 + 모든 P값 표시"""
    plt.figure(figsize=(10, 6))
    
    # 히스토그램
    plt.hist(res["elapsed_days"], bins=40, alpha=0.7, color='skyblue', edgecolor='black')
    
    # P값 수직선
    for p in pvals:
        val = np.percentile(res["elapsed_days"], p)
        color, lw, ls = get_pvalue_style(p, pvals)
        plt.axvline(val, color=color, linewidth=lw, linestyle=ls, 
                   label=f'P{p} = {val:.1f}일')
    
    plt.xlabel("캠페인 기간 (일)", fontsize=12)
    plt.ylabel("빈도", fontsize=12)
    plt.title("캠페인 기간 분포 및 백분위수", fontsize=14, fontweight='bold')
    plt.legend(loc='upper right')
    plt.grid(alpha=0.3)
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        plt.close()
    else:
        plt.show()


def plot_timeline(res: pd.DataFrame, n_samples: int = 5, save_path: str = None):
    """작업/대기 타임라인 (상위 N개 샘플)"""
    samples = res.nlargest(n_samples, 'elapsed_days')
    
    fig, axes = plt.subplots(n_samples, 1, figsize=(12, n_samples*1.2), sharex=True)
    if n_samples == 1:
        axes = [axes]
    
    for idx, (i, row) in enumerate(samples.iterrows()):
        ax = axes[idx]
        work_days = row['work_hours'] / 24
        wait_days = row['wait_hours'] / 24
        total_days = row['elapsed_days']
        
        # 막대 그래프
        ax.barh(0, work_days, color='green', alpha=0.7, label='작업')
        ax.barh(0, wait_days, left=work_days, color='red', alpha=0.7, label='대기')
        
        # 텍스트
        work_pct = work_days / total_days * 100
        wait_pct = wait_days / total_days * 100
        
        ax.text(work_days/2, 0, f'{work_days:.1f}일\n({work_pct:.0f}%)', 
               ha='center', va='center', fontsize=9, fontweight='bold')
        ax.text(work_days + wait_days/2, 0, f'{wait_days:.1f}일\n({wait_pct:.0f}%)', 
               ha='center', va='center', fontsize=9, fontweight='bold')
        
        ax.set_yticks([])
        ax.set_xlim(0, total_days * 1.05)
        ax.set_title(f'시뮬레이션 #{row["sim"]+1} (총 {total_days:.1f}일)', 
                    fontsize=10, loc='left')
        
        if idx == 0:
            ax.legend(loc='upper right', fontsize=9)
        
        ax.grid(axis='x', alpha=0.3)
    
    axes[-1].set_xlabel("기간 (일)", fontsize=11)
    fig.suptitle(f'작업/대기 타임라인 (상위 {n_samples}개 샘플)', 
                fontsize=13, fontweight='bold', y=0.995)
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        plt.close()
    else:
        plt.show()


def plot_cdf(res: pd.DataFrame, save_path: str = None):
    """누적 분포 함수 (CDF)"""
    sorted_days = np.sort(res["elapsed_days"])
    cdf = np.arange(1, len(sorted_days)+1) / len(sorted_days) * 100
    
    plt.figure(figsize=(10, 6))
    plt.plot(sorted_days, cdf, linewidth=2, color='navy')
    plt.xlabel("캠페인 기간 (일)", fontsize=12)
    plt.ylabel("누적 확률 (%)", fontsize=12)
    plt.title("누적 분포 함수 (CDF)", fontsize=14, fontweight='bold')
    plt.grid(alpha=0.3)
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        plt.close()
    else:
        plt.show()


def plot_work_wait_scatter(res: pd.DataFrame, save_path: str = None):
    """작업 vs 대기 산점도"""
    plt.figure(figsize=(8, 6))
    plt.scatter(res["work_hours"]/24, res["wait_hours"]/24, alpha=0.5, s=20)
    plt.xlabel("작업 시간 (일)", fontsize=12)
    plt.ylabel("대기 시간 (일)", fontsize=12)
    plt.title("작업 시간 vs 대기 시간", fontsize=14, fontweight='bold')
    plt.grid(alpha=0.3)
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        plt.close()
    else:
        plt.show()


def plot_calendar_availability(df: pd.DataFrame, calendar_fn, save_path: str = None):
    """업무 시간 제약 시각화"""
    if calendar_fn is None:
        return
    
    # 시간대별 작업 가능 비율
    hours = df.index.hour.value_counts().sort_index()
    mask = calendar_fn(df.index)
    available_hours = df.index[mask].hour.value_counts().sort_index()
    
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
    
    # 시간대별 막대 그래프
    x = list(range(24))
    total_per_hour = [hours.get(h, 0) for h in x]
    avail_per_hour = [available_hours.get(h, 0) for h in x]
    ratio = [a/t*100 if t > 0 else 0 for a, t in zip(avail_per_hour, total_per_hour)]
    
    ax1.bar(x, ratio, color='green', alpha=0.7)
    ax1.set_xlabel("시간 (시)", fontsize=11)
    ax1.set_ylabel("작업 가능 비율 (%)", fontsize=11)
    ax1.set_title("시간대별 작업 가능 비율", fontsize=12, fontweight='bold')
    ax1.set_xticks(range(0, 24, 2))
    ax1.grid(alpha=0.3)
    
    # 전체 가용 시간 파이 차트
    total_hours = len(df)
    avail_hours = mask.sum()
    blocked_hours = total_hours - avail_hours
    
    ax2.pie([avail_hours, blocked_hours], 
           labels=['작업 가능', '작업 불가'],
           autopct='%1.1f%%',
           colors=['green', 'lightgray'],
           startangle=90)
    ax2.set_title("전체 가용 시간 비율", fontsize=12, fontweight='bold')
    
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        plt.close()
    else:
        plt.show()


# ============ Excel 리포트 ============
def generate_excel_report(res: pd.DataFrame, summary: pd.DataFrame, 
                         config: dict, save_path: str):
    """Excel 리포트 생성"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logging.warning("openpyxl 미설치. Excel 리포트 생성 불가")
        return
    
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        # Sheet 1: 요약 통계
        summary.to_excel(writer, sheet_name='요약', index=False)
        
        # Sheet 2: 전체 결과
        res.to_excel(writer, sheet_name='전체결과', index=False)
        
        # Sheet 3: 작업 정보
        tasks_info = pd.DataFrame(config.get('tasks', []))
        if not tasks_info.empty:
            tasks_info.to_excel(writer, sheet_name='작업정보', index=False)
        
        # 서식 적용
        wb = writer.book
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 헤더 스타일
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    logging.info(f"Excel 리포트 저장: {save_path}")


# ============ 배치 실행 함수 ============
def batch_run_multiple_files(csv_files: List[str], config: dict, csv_type: str = 'general'):
    """다중 파일 배치 분석"""
    all_results = {}
    
    for csv_path in csv_files:
        logging.info(f"\n{'='*60}")
        logging.info(f"분석 시작: {os.path.basename(csv_path)}")
        logging.info(f"{'='*60}")
        
        try:
            # CSV 로드
            df = load_csv(csv_path, csv_type)
            
            # 검증
            valid, msg = validate_metocean(df)
            if not valid:
                logging.error(f"검증 실패: {msg}")
                continue
            
            # 시간 간격
            interval_min = get_time_interval_minutes(df)
            logging.info(f"✓ 데이터 로드 완료: {len(df)}개, {interval_min}분 간격")
            
            # 시뮬레이션
            tasks = [Task(**t) for t in config['tasks']]
            res = simulate_campaign(
                df, tasks,
                n_sims=config.get('n_sims', 1000),
                start_month=config.get('start_month'),
                split_mode=config.get('split_mode', False),
                time_interval_min=interval_min,
                na_handling=config.get('na_handling', 'permissive')
            )
            
            # 요약
            summary = summarize_pxx(res, config.get('pvals', [50, 75, 90]))
            
            file_name = os.path.basename(csv_path).replace('.csv', '')
            all_results[file_name] = {
                'results': res,
                'summary': summary
            }
            
            logging.info("\n결과 요약:")
            print(summary.to_string(index=False))
            
        except Exception as e:
            logging.error(f"분석 실패: {csv_path} - {str(e)}")
            continue
    
    return all_results


def plot_comparison(all_results: dict, save_dir: str):
    """지점 비교 차트"""
    # 박스플롯
    plt.figure(figsize=(12, 6))
    data = [res['results']['elapsed_days'] for res in all_results.values()]
    labels = list(all_results.keys())
    
    plt.boxplot(data, labels=labels)
    plt.ylabel("캠페인 기간 (일)", fontsize=12)
    plt.title("지점별 캠페인 기간 비교", fontsize=14, fontweight='bold')
    plt.xticks(rotation=45, ha='right')
    plt.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'comparison_boxplot.png'), dpi=150)
    plt.close()
    
    # 요약 테이블
    comparison_df = pd.DataFrame({
        name: res['summary']['Value_days'].values
        for name, res in all_results.items()
    }, index=all_results[list(all_results.keys())[0]]['summary']['Metric'])
    
    comparison_df.to_csv(os.path.join(save_dir, 'comparison.csv'))
    logging.info(f"비교 결과 저장: {save_dir}")


# ============ 최적 시작월 분석 ============
def analyze_optimal_start_month(df: pd.DataFrame, config: dict, 
                                time_interval_min: int) -> pd.DataFrame:
    """월별 P90 비교 및 최적 시작월 추천"""
    logging.info("\n최적 시작월 분석 중...")
    
    tasks = [Task(**t) for t in config['tasks']]
    monthly_results = []
    
    for month in range(1, 13):
        print(f"  {month}월 분석 중...")
        
        res = simulate_campaign(
            df, tasks,
            n_sims=config.get('n_sims', 500),  # 빠른 분석
            start_month=month,
            split_mode=config.get('split_mode', False),
            time_interval_min=time_interval_min,
            na_handling=config.get('na_handling', 'permissive'),
            show_progress=False
        )
        
        p90 = np.percentile(res['elapsed_days'], 90)
        mean = res['elapsed_days'].mean()
        
        monthly_results.append({
            'Month': month,
            'P90_days': p90,
            'Mean_days': mean
        })
    
    result_df = pd.DataFrame(monthly_results)
    
    # 최적 월
    optimal_month = result_df.loc[result_df['P90_days'].idxmin(), 'Month']
    min_p90 = result_df['P90_days'].min()
    
    logging.info(f"\n✓ 최적 시작월: {int(optimal_month)}월 (P90 = {min_p90:.1f}일)")
    
    # 차트
    plt.figure(figsize=(10, 6))
    plt.plot(result_df['Month'], result_df['P90_days'], marker='o', linewidth=2, label='P90')
    plt.plot(result_df['Month'], result_df['Mean_days'], marker='s', linewidth=2, label='Mean', alpha=0.7)
    plt.axvline(optimal_month, color='red', linestyle='--', alpha=0.5, label=f'최적: {int(optimal_month)}월')
    
    plt.xlabel("시작 월", fontsize=12)
    plt.ylabel("캠페인 기간 (일)", fontsize=12)
    plt.title("월별 캠페인 기간 분석", fontsize=14, fontweight='bold')
    plt.xticks(range(1, 13))
    plt.legend()
    plt.grid(alpha=0.3)
    plt.tight_layout()
    plt.show()
    
    return result_df


# ============ 사용자 입력 함수 ============
def ask(msg, default=None):
    s = input(f"{msg}" + (f" [기본값: {default}]" if default else "") + ": ").strip()
    return s if s else (default if default else "")

def ask_int(msg, default=None):
    while True:
        s = ask(msg, default)
        try:
            return int(s)
        except:
            print(" → 정수로 입력해주세요")

def ask_yesno(msg, default="y"):
    s = input(f"{msg} [기본값: {default}] : ").strip().lower()
    if s == "":
        s = default
    return s.startswith("y")


def read_tasks_json():
    """JSON 형식으로 작업 목록 입력"""
    print("\nJSON 붙여넣기 (빈 줄로 종료):")
    lines = []
    while True:
        line = input()
        if line.strip() == "":
            break
        lines.append(line)
    raw = "\n".join(lines)
    data = json.loads(raw)
    return data.get("tasks", []), data


# ============ 메인 실행 함수 ============
def run_simulation(df: pd.DataFrame, config: dict, output_dir: str = None):
    """단일 파일 시뮬레이션 실행"""
    # 검증
    valid, msg = validate_metocean(df)
    if not valid:
        logging.error(f"✗ 데이터 검증 실패: {msg}")
        return False
    
    logging.info(f"✓ 데이터 검증 통과")
    
    # 시간 간격
    interval_min = get_time_interval_minutes(df)
    logging.info(f"✓ 시간 간격: {interval_min}분")
    
    # 작업 정보
    tasks = [Task(**t) for t in config['tasks']]
    logging.info(f"✓ 작업 수: {len(tasks)}개")
    
    # 캘린더 설정
    calendar_mode = config.get('calendar', ['all'])[0]
    cal_fn = None
    
    if calendar_mode == 'custom':
        hours = config['calendar'][2]
        sh, eh = map(int, hours.split('-'))
        def cal_fn(index):
            hrs = index.hour
            return (hrs >= sh) & (hrs < eh)
        logging.info(f"✓ 업무 시간: {sh}시~{eh}시")
    else:
        logging.info(f"✓ 업무 시간: 24시간")
    
    # 시뮬레이션
    logging.info(f"\n시뮬레이션 시작 (n_sims={config.get('n_sims', 1000)})...")
    
    res = simulate_campaign(
        df, tasks,
        n_sims=config.get('n_sims', 1000),
        start_month=config.get('start_month'),
        calendar_mask_fn=cal_fn,
        split_mode=config.get('split_mode', False),
        time_interval_min=interval_min,
        na_handling=config.get('na_handling', 'permissive')
    )
    
    # 요약
    pvals = config.get('pvals', [50, 75, 90])
    summary = summarize_pxx(res, pvals)
    
    logging.info("\n" + "="*50)
    logging.info("결과 요약")
    logging.info("="*50)
    print(summary.to_string(index=False))
    
    # 저장
    if output_dir is None:
        if not ask_yesno("\n결과를 저장하시겠습니까?", "y"):
            return True
        output_dir = ask("저장 폴더", "./pxx_outputs")
    
    os.makedirs(output_dir, exist_ok=True)
    
    # CSV
    res.to_csv(os.path.join(output_dir, 'sim_results.csv'), index=False)
    summary.to_csv(os.path.join(output_dir, 'summary.csv'), index=False)
    
    # 차트
    plot_histogram_with_pvals(res, pvals, os.path.join(output_dir, 'histogram_pvals.png'))
    plot_timeline(res, 5, os.path.join(output_dir, 'timeline.png'))
    plot_cdf(res, os.path.join(output_dir, 'cdf.png'))
    plot_work_wait_scatter(res, os.path.join(output_dir, 'work_vs_wait.png'))
    
    if cal_fn:
        plot_calendar_availability(df, cal_fn, os.path.join(output_dir, 'calendar.png'))
    
    # Excel
    generate_excel_report(res, summary, config, os.path.join(output_dir, 'report.xlsx'))
    
    logging.info(f"\n✓ 결과 저장 완료: {output_dir}")
    return True


# ============ 메인 루프 ============
def main():
    print("="*60)
    print("Marine Pxx Simulator V5")
    print("="*60)
    
    last_df = None
    last_config = None
    last_csv_type = 'general'
    
    while True:
        print("\n" + "-"*60)
        print("메뉴")
        print("-"*60)
        print("[1] 단일 파일 분석 (일반 CSV)")
        print("    → 일반 기상 데이터")
        print("[2] 단일 파일 분석 (hindcast CSV)")
        print("    → ERA5 30년 hindcast 데이터")
        print("[3] 다중 파일 배치 분석")
        print("    → 여러 지점 동시 비교")
        print("[4] 최적 시작월 분석")
        print("    → 월별 P90 비교 (메뉴 1/2 먼저 실행)")
        print("[5] 설정 파일로 배치 실행")
        print("    → JSON 설정 파일 사용")
        print("[6] 이전 설정 재사용")
        print("    → 마지막 실행 설정 불러오기")
        print("[7] 종료")
        print("-"*60)
        
        choice = ask_int("선택", 1)
        
        # [1] 일반 CSV
        if choice == 1:
            csv_path = ask("CSV 파일 경로")
            df = load_csv(csv_path, 'general')
            
            tasks_list, config = read_tasks_json()
            config['tasks'] = tasks_list
            
            last_df = df
            last_config = config
            last_csv_type = 'general'
            
            run_simulation(df, config)
        
        # [2] hindcast CSV
        elif choice == 2:
            csv_path = ask("hindcast CSV 파일 경로")
            
            # 날짜 구간 설정 (선택 사항)
            use_date_filter = ask_yesno("\n날짜 구간을 설정하시겠습니까?", "n")
            start_date = None
            end_date = None
            
            if use_date_filter:
                start_date = ask("시작 날짜", "1990-01-01")
                end_date = ask("종료 날짜", "2019-12-31")
                
                # 빈 문자열이면 None으로 처리
                start_date = start_date if start_date else None
                end_date = end_date if end_date else None
            
            df = load_csv(csv_path, 'hindcast', start_date, end_date)
            
            tasks_list, config = read_tasks_json()
            config['tasks'] = tasks_list
            
            last_df = df
            last_config = config
            last_csv_type = 'hindcast'
            
            run_simulation(df, config)
        
        # [3] 배치 분석
        elif choice == 3:
            csv_type = ask("CSV 타입 (general/hindcast)", "general")
            
            print("\nCSV 파일 경로 입력 (빈 줄로 종료):")
            csv_files = []
            while True:
                path = input("파일: ").strip()
                if not path:
                    break
                csv_files.append(path)
            
            if not csv_files:
                print("파일이 입력되지 않았습니다.")
                continue
            
            tasks_list, config = read_tasks_json()
            config['tasks'] = tasks_list
            
            results = batch_run_multiple_files(csv_files, config, csv_type)
            
            if results:
                save_dir = ask("결과 저장 폴더", "./batch_outputs")
                os.makedirs(save_dir, exist_ok=True)
                plot_comparison(results, save_dir)
        
        # [4] 최적 시작월
        elif choice == 4:
            if last_df is None or last_config is None:
                print("먼저 메뉴 1 또는 2를 실행해주세요.")
                continue
            
            interval_min = get_time_interval_minutes(last_df)
            result = analyze_optimal_start_month(last_df, last_config, interval_min)
            print("\n월별 분석 결과:")
            print(result.to_string(index=False))
        
        # [5] 설정 파일 배치
        elif choice == 5:
            config_path = ask("설정 JSON 파일 경로")
            with open(config_path, 'r', encoding='utf-8') as f:
                batch_config = json.load(f)
            
            csv_files = batch_config.get('csv_files', [])
            config = batch_config.get('config', {})
            csv_type = batch_config.get('csv_type', 'general')
            
            results = batch_run_multiple_files(csv_files, config, csv_type)
            
            if results:
                save_dir = batch_config.get('output_dir', './batch_outputs')
                os.makedirs(save_dir, exist_ok=True)
                plot_comparison(results, save_dir)
        
        # [6] 이전 설정 재사용
        elif choice == 6:
            if last_df is None or last_config is None:
                print("이전 설정이 없습니다.")
                continue
            
            run_simulation(last_df, last_config)
        
        # [7] 종료
        else:
            print("프로그램을 종료합니다.")
            break


if __name__ == "__main__":
    main()
